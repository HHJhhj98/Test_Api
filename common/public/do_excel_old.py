# -*- coding：utf-8 -*-
# @Time ：2022/1/3 10:35
# @Authon :hhj
# @Annotation:
# @File : do_excel_old.py

from openpyxl import load_workbook
import os, sys

from openpyxl.styles import PatternFill

from common.public.do_config import DoConfig
from common.public.get_data import GetData
from common.public.project_path import *
from common.public.my_log import MyLog

my_logger = MyLog()


class DoExcel:

    def read_excel(self, filename):
        wb = load_workbook(filename)
        test_data = []
        tel = getattr(GetData, 'NoRegTel')
        normal_tel = getattr(GetData, 'NORMAL_TEL')
        loan_tel = getattr(GetData, 'LOAN_TEL')
        admin_tel = getattr(GetData, 'ADMIN_TEL')
        print(tel)
        mode = eval(DoConfig().read_config(conf_path, 'MODE', 'mode'))
        for key in mode:  # 遍历存在配置文件里面的字典
            sheet = wb[key]

            if mode[key] == 'all':
                for i in range(2, sheet.max_row + 1):
                    case_data = {}
                    case_data['case_id'] = sheet.cell(row=i, column=1).value
                    case_data['module'] = sheet.cell(row=i, column=2).value
                    case_data['title'] = sheet.cell(row=i, column=3).value
                    case_data['url'] = sheet.cell(row=i, column=4).value
                    case_data['method'] = sheet.cell(row=i, column=5).value
                    case_data['data'] = sheet.cell(row=i, column=6).value
                    # if sheet.cell(i, 6).value.find('${tel_1}') != -1:
                    #     case_data['data'] = eval(sheet.cell(row=i, column=6).value.replace('${tel_1}', str(tel)))
                    # elif sheet.cell(i, 6).value.find('${tel_2}') != -1:
                    #     case_data['data'] = eval(sheet.cell(row=i, column=6).value.replace('${tel_2}', str(tel + 1)))
                    # elif sheet.cell(i, 6).value.find('${tel_3}') != -1:
                    #     case_data['data'] = eval(sheet.cell(row=i, column=6).value.replace('${tel_3}', str(tel + 2)))
                    # else:
                    #     case_data['data'] = eval(sheet.cell(row=i, column=6).value)
                    case_data['headers'] = sheet.cell(row=i, column=7).value
                    case_data['expected'] = eval(sheet.cell(row=i, column=8).value)
                    case_data['check_sql_list'] = eval(sheet.cell(row=i, column=10).value)
                    # print(case_data['data'])
                    test_data.append(case_data)

            else:
                # print(mode[key], type(mode[key]))
                for case_id in mode[key]:
                    case_data = {}
                    case_data['case_id'] = sheet.cell(row=case_id + 1, column=1).value
                    case_data['module'] = sheet.cell(row=case_id + 1, column=2).value
                    case_data['title'] = sheet.cell(row=case_id + 1, column=3).value
                    case_data['url'] = sheet.cell(row=case_id + 1, column=4).value
                    case_data['method'] = sheet.cell(row=case_id + 1, column=5).value
                    case_data['data'] = sheet.cell(row=case_id + 1, column=6).value
                    # if sheet.cell(i, 6).value.find('${tel_1}') != -1:
                    #     case_data['data'] = eval(
                    #         sheet.cell(row=case_id + 1, column=6).value.replace('${tel_1}', str(tel)))
                    # elif sheet.cell(i, 6).value.find('${tel_2}') != -1:
                    #     case_data['data'] = eval(
                    #         sheet.cell(row=case_id + 1, column=6).value.replace('${tel_2}', str(tel + 1)))
                    # elif sheet.cell(i, 6).value.find('${tel_3}') != -1:
                    #     case_data['data'] = eval(
                    #         sheet.cell(row=case_id + 1, column=6).value.replace('${tel_3}', str(tel + 2)))
                    # else:
                    #     case_data['data'] = eval(sheet.cell(row=case_id + 1, column=6).value)
                    case_data['headers'] = sheet.cell(row=case_id + 1, column=7).value
                    case_data['expected'] = eval(sheet.cell(row=case_id + 1, column=8).value)
                    case_data['check_sql_list'] = eval(sheet.cell(row=case_id + 1, column=10).value)

                    # dict(
                    #     case_id=sheet.cell(row=case_id+1, column=1).value,
                    #     module=sheet.cell(row=case_id+1, column=2).value,
                    #     title=sheet.cell(row=case_id+1, column=3).value,
                    #     url=sheet.cell(row=case_id+1, column=4).value,
                    #     method=sheet.cell(row=case_id+1, column=5).value,
                    #     data=eval(sheet.cell(row=case_id+1, column=6).value),
                    #     headers=eval(sheet.cell(row=case_id+1, column=7).value),
                    #     expected=eval(sheet.cell(row=case_id+1, column=8).value)
                    # )
                    test_data.append(case_data)

            self.update_tel(tel + 3, filename, 'init')
        for item in test_data:
            if item['data'].find('${tel_1}') != -1:
                item['data'] = item['data'].replace('${tel_1}', str(tel))
            elif item['data'].find('${tel_2}') != -1:
                item['data'] = item['data'].replace('${tel_2}', str(tel + 1))
            elif item['data'].find('${tel_3}') != -1:
                item['data'] = item['data'].replace('${tel_3}', str(tel + 2))
            elif item['data'].find('${normal_tel}') != -1:
                item['data'] = item['data'].replace('${normal_tel}', str(normal_tel))
            elif item['data'].find('${loan_tel}') != -1:
                item['data'] = item['data'].replace('${loan_tel}', str(loan_tel))
            elif item['data'].find('${admin_tel}') != -1:
                item['data'] = item['data'].replace('${admin_tel}', str(admin_tel))
            else:
                item['data'] = item['data']
            # print(item)
        my_logger.info("excel:{}".format(test_data))
        return test_data

    # def write_excel(self, filename, sheetname, case_id, result, test_result):
    #     wb = load_workbook(filename)
    #     sheet = wb[sheetname]
    #     # 设置样式，并且加载到对应单元格
    #     if test_result == 'pass':
    #         fill = PatternFill("solid", fgColor="009100")
    #     else:
    #         fill = PatternFill("solid", fgColor="EA0000")
    #
    #     sheet.cell(case_id + 1, 9).value = result
    #     sheet.cell(case_id + 1, 12).value = test_result
    #     sheet.cell(case_id + 1, 12).fill = fill
    #     wb.save(filename)

    def write_excel(self, filename, sheetname, row, col, result=None, test_result=None):
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        # 设置样式，并且加载到对应单元格
        if test_result == 'pass':
            fill = PatternFill("solid", fgColor="009100")
            sheet.cell(row, col).fill = fill
            sheet.cell(row, col).value = test_result
        elif test_result == 'fail':
            fill = PatternFill("solid", fgColor="EA0000")
            sheet.cell(row, col).fill = fill
            sheet.cell(row, col).value = test_result

        sheet.cell(row, col).value = result
        wb.save(filename)

    def write_sql_excel(self, filename, sheetname, case_id, test_result):
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        # # 设置样式，并且加载到对应单元格
        # if test_result == 'pass':
        #     fill = PatternFill("solid", fgColor="009100")
        # else:
        #     fill = PatternFill("solid", fgColor="EA0000")

        sheet.cell(case_id + 1, 11).value = test_result
        wb.save(filename)

    def update_tel(self, tel, file_name, sheet_name):
        wb = load_workbook(file_name)
        wb[sheet_name].cell(2, 1).value = tel
        wb.save(file_name)


if __name__ == '__main__':
    print(DoExcel().read_excel(test_data_path))
